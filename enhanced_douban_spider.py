#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
豆瓣书单爬虫 - 增强版
支持爬取全部书籍并生成带封面的Excel文件
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import json
import re
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import io

def crawl_all_douban_books(doulist_url, max_pages=20):
    """
    爬取豆瓣书单中的所有书籍信息
    
    Args:
        doulist_url: 豆瓣书单URL
        max_pages: 最大爬取页数（每页25本，20页=500本）
    
    Returns:
        list: 包含书籍信息的字典列表
    """
    
    # 设置请求头
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    }
    
    books_data = []
    session = requests.Session()
    session.headers.update(headers)
    
    page = 0
    start = 0
    
    print(f"开始爬取豆瓣书单: {doulist_url}")
    print(f"计划爬取最多 {max_pages} 页，约 {max_pages * 25} 本书籍")
    
    while page < max_pages:
        # 构建当前页面URL
        if start == 0:
            current_url = doulist_url
        else:
            # 替换URL中的start参数
            current_url = re.sub(r'start=\d+', f'start={start}', doulist_url)
        
        try:
            print(f"正在爬取第 {page + 1} 页...")
            response = session.get(current_url, timeout=10)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            soup = BeautifulSoup(response.text, 'html.parser')
            items = soup.find_all('div', class_='doulist-item')
            
            if not items:
                print(f"第 {page + 1} 页没有找到书籍，可能已到最后一页")
                break
            
            print(f"第 {page + 1} 页找到 {len(items)} 本书籍")
            
            # 解析每本书的信息
            page_books = 0
            for i, item in enumerate(items):
                book_info = parse_single_book(item)
                if book_info['书名']:
                    books_data.append(book_info)
                    page_books += 1
                    print(f"  ✓ {book_info['书名']} - {book_info['评分']}")
            
            print(f"第 {page + 1} 页成功解析 {page_books} 本书籍")
            print(f"累计已爬取 {len(books_data)} 本书籍")
            
            # 检查是否还有下一页
            next_page = soup.find('span', class_='next')
            if not next_page or not next_page.find('a'):
                print("没有找到下一页链接，爬取完成")
                break
            
            # 准备下一页
            start += 25
            page += 1
            
            # 添加延时
            print("等待 2 秒...")
            time.sleep(2)
            
        except requests.RequestException as e:
            print(f"请求第 {page + 1} 页时出错: {e}")
            break
        except Exception as e:
            print(f"处理第 {page + 1} 页时出错: {e}")
            break
    
    print(f"\n爬取完成！共获取 {len(books_data)} 本书籍")
    return books_data

def parse_single_book(item):
    """解析单个书籍的信息"""
    book_info = {
        '书名': '',
        '作者': '',
        '出版社': '',
        '评分': '',
        '封面链接': '',
        '书籍链接': ''
    }
    
    try:
        # 获取书名和链接
        title_div = item.find('div', class_='title')
        if title_div:
            title_link = title_div.find('a')
            if title_link:
                book_info['书名'] = title_link.get_text(strip=True)
                book_info['书籍链接'] = title_link.get('href', '')
        
        # 获取封面
        post_div = item.find('div', class_='post')
        if post_div:
            img = post_div.find('img')
            if img:
                book_info['封面链接'] = img.get('src', '')
        
        # 获取评分
        rating_div = item.find('div', class_='rating')
        if rating_div:
            rating_span = rating_div.find('span', class_='rating_nums')
            if rating_span:
                book_info['评分'] = rating_span.get_text(strip=True)
        
        # 获取详细信息
        abstract_div = item.find('div', class_='abstract')
        if abstract_div:
            abstract_text = abstract_div.get_text(strip=True)
            
            # 使用正则表达式提取作者和出版社
            author_match = re.search(r'作者[:：]\s*(.+?)(?=出版社|出版年|$)', abstract_text)
            if author_match:
                book_info['作者'] = author_match.group(1).strip()
            
            publisher_match = re.search(r'出版社[:：]\s*(.+?)(?=出版年|$)', abstract_text)
            if publisher_match:
                book_info['出版社'] = publisher_match.group(1).strip()
    
    except Exception as e:
        print(f"解析书籍信息时出错: {e}")
    
    return book_info

def download_image(url, filename):
    """下载图片到本地"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        with open(filename, 'wb') as f:
            f.write(response.content)
        return True
    except Exception as e:
        print(f"下载图片失败 {url}: {e}")
        return False

def create_excel_with_covers(books_data, excel_file='douban_books_with_covers.xlsx'):
    """创建带封面的Excel文件"""
    print(f"\n开始创建Excel文件: {excel_file}")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "豆瓣书单"
    
    # 设置列标题
    headers = ['序号', '书名', '作者', '出版社', '评分', '书籍链接', '封面']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 创建封面图片文件夹
    covers_dir = 'book_covers'
    if not os.path.exists(covers_dir):
        os.makedirs(covers_dir)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 30  # 书名
    ws.column_dimensions['C'].width = 25  # 作者
    ws.column_dimensions['D'].width = 20  # 出版社
    ws.column_dimensions['E'].width = 8   # 评分
    ws.column_dimensions['F'].width = 50  # 链接
    ws.column_dimensions['G'].width = 15  # 封面
    
    # 设置行高
    ws.row_dimensions[1].height = 20
    
    print(f"开始处理 {len(books_data)} 本书籍...")
    
    for i, book in enumerate(books_data, 2):
        print(f"处理第 {i-1} 本书: {book['书名']}")
        
        # 填充基本信息
        ws.cell(row=i, column=1, value=i-1)  # 序号
        ws.cell(row=i, column=2, value=book['书名'])  # 书名
        ws.cell(row=i, column=3, value=book['作者'])  # 作者
        ws.cell(row=i, column=4, value=book['出版社'])  # 出版社
        ws.cell(row=i, column=5, value=book['评分'])  # 评分
        ws.cell(row=i, column=6, value=book['书籍链接'])  # 链接
        
        # 下载并插入封面
        if book['封面链接']:
            try:
                # 生成文件名
                safe_title = re.sub(r'[^\w\s-]', '', book['书名'])[:20]  # 限制文件名长度
                image_filename = f"{covers_dir}/cover_{i-1}_{safe_title}.jpg"
                
                # 下载图片
                if download_image(book['封面链接'], image_filename):
                    # 调整图片大小
                    img = PILImage.open(image_filename)
                    img.thumbnail((100, 140), PILImage.Resampling.LANCZOS)
                    img.save(image_filename, 'JPEG', quality=85)
                    
                    # 插入到Excel
                    excel_img = Image(image_filename)
                    excel_img.width = 80
                    excel_img.height = 112
                    
                    # 设置图片位置
                    cell_ref = f'G{i}'
                    ws.add_image(excel_img, cell_ref)
                    
                    # 设置行高以适应图片
                    ws.row_dimensions[i].height = 90
                    
                    print(f"  ✓ 封面已添加")
                else:
                    ws.cell(row=i, column=7, value="封面下载失败")
                    print(f"  ✗ 封面下载失败")
                    
            except Exception as e:
                ws.cell(row=i, column=7, value=f"封面处理失败: {str(e)}")
                print(f"  ✗ 封面处理失败: {e}")
        else:
            ws.cell(row=i, column=7, value="无封面")
            print(f"  - 无封面链接")
        
        # 设置文本对齐
        for col in range(1, 7):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')
    
    # 保存Excel文件
    wb.save(excel_file)
    print(f"\nExcel文件已保存: {excel_file}")
    print(f"封面图片已保存到: {covers_dir}/ 文件夹")
    
    return excel_file

def save_data(books_data, csv_file='douban_books_all.csv', json_file='douban_books_all.json'):
    """保存数据到文件"""
    if not books_data:
        print("没有数据可保存")
        return
    
    # 保存为CSV
    df = pd.DataFrame(books_data)
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"数据已保存到 {csv_file}")
    
    # 保存为JSON
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(books_data, f, ensure_ascii=False, indent=2)
    print(f"数据已保存到 {json_file}")

def print_statistics(books_data):
    """打印统计信息"""
    if not books_data:
        print("没有数据可统计")
        return
    
    print(f"\n=== 统计信息 ===")
    print(f"总书籍数量: {len(books_data)}")
    
    # 有评分的书籍
    rated_books = [book for book in books_data if book['评分']]
    print(f"有评分的书籍: {len(rated_books)}")
    
    # 评分统计
    if rated_books:
        scores = [float(book['评分']) for book in rated_books if book['评分']]
        if scores:
            print(f"平均评分: {sum(scores)/len(scores):.2f}")
            print(f"最高评分: {max(scores)}")
            print(f"最低评分: {min(scores)}")
    
    # 出版社统计
    publishers = [book['出版社'] for book in books_data if book['出版社']]
    if publishers:
        from collections import Counter
        publisher_counts = Counter(publishers)
        print(f"\n出版社分布 (前5名):")
        for publisher, count in publisher_counts.most_common(5):
            print(f"  {publisher}: {count}本")

def main():
    """主函数"""
    # 目标豆瓣书单URL
    doulist_url = "https://www.douban.com/doulist/45298673/?start=0&sort=seq&playable=0&sub_type="
    
    try:
        # 开始爬取（爬取20页，约500本书）
        books_data = crawl_all_douban_books(doulist_url, max_pages=20)
        
        if books_data:
            # 保存数据
            save_data(books_data)
            
            # 创建带封面的Excel文件
            excel_file = create_excel_with_covers(books_data)
            
            # 统计信息
            print_statistics(books_data)
            
            print(f"\n=== 完成 ===")
            print(f"✓ 成功爬取 {len(books_data)} 本书籍")
            print(f"✓ 数据已保存为CSV和JSON格式")
            print(f"✓ 带封面的Excel文件已生成: {excel_file}")
            print(f"✓ 封面图片已下载到 book_covers/ 文件夹")
            
        else:
            print("没有爬取到任何数据")
    
    except KeyboardInterrupt:
        print("\n用户中断爬取")
    except Exception as e:
        print(f"程序执行出错: {e}")

if __name__ == "__main__":
    main()
