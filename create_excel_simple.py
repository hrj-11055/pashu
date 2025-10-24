#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化Excel生成器 - 为已爬取的数据生成带封面的Excel文件
"""

import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from PIL import Image as PILImage

def create_excel_from_data():
    """从已爬取的数据创建Excel文件"""
    
    # 读取JSON数据
    print("正在读取数据文件...")
    with open('douban_books_all.json', 'r', encoding='utf-8') as f:
        books_data = json.load(f)
    
    print(f"读取到 {len(books_data)} 本书籍数据")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "豆瓣书单"
    
    # 设置列标题
    headers = ['序号', '书名', '作者', '出版社', '评分', '书籍链接', '封面状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 30  # 书名
    ws.column_dimensions['C'].width = 25  # 作者
    ws.column_dimensions['D'].width = 20  # 出版社
    ws.column_dimensions['E'].width = 8   # 评分
    ws.column_dimensions['F'].width = 50  # 链接
    ws.column_dimensions['G'].width = 15  # 封面状态
    
    # 设置标题行高
    ws.row_dimensions[1].height = 20
    
    print("开始处理书籍数据...")
    
    # 获取所有封面文件
    cover_files = os.listdir('book_covers')
    print(f"找到 {len(cover_files)} 个封面文件")
    
    # 处理每本书
    for i, book in enumerate(books_data, 2):
        if i % 50 == 0:  # 每50本书显示一次进度
            print(f"已处理 {i-1} 本书籍...")
        
        # 填充基本信息
        ws.cell(row=i, column=1, value=i-1)  # 序号
        ws.cell(row=i, column=2, value=book['书名'])  # 书名
        ws.cell(row=i, column=3, value=book['作者'])  # 作者
        ws.cell(row=i, column=4, value=book['出版社'])  # 出版社
        ws.cell(row=i, column=5, value=book['评分'])  # 评分
        ws.cell(row=i, column=6, value=book['书籍链接'])  # 链接
        
        # 检查封面状态
        cover_status = "无封面"
        if book['封面链接']:
            # 查找对应的封面文件
            cover_files_for_book = [f for f in cover_files if f.startswith(f"cover_{i-1}_")]
            if cover_files_for_book:
                cover_status = f"已下载 ({cover_files_for_book[0]})"
            else:
                cover_status = "封面链接存在但未下载"
        
        ws.cell(row=i, column=7, value=cover_status)
        
        # 设置文本对齐
        for col in range(1, 8):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')
    
    # 保存Excel文件
    excel_file = 'douban_books_with_covers.xlsx'
    wb.save(excel_file)
    print(f"\nExcel文件已保存: {excel_file}")
    
    return excel_file

def create_excel_with_images():
    """创建带图片的Excel文件（处理前50本书）"""
    
    # 读取JSON数据
    print("正在读取数据文件...")
    with open('douban_books_all.json', 'r', encoding='utf-8') as f:
        books_data = json.load(f)
    
    # 只处理前50本书（避免文件过大）
    books_data = books_data[:50]
    print(f"处理前 {len(books_data)} 本书籍")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "豆瓣书单(带封面)"
    
    # 设置列标题
    headers = ['序号', '书名', '作者', '出版社', '评分', '书籍链接', '封面']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 30  # 书名
    ws.column_dimensions['C'].width = 25  # 作者
    ws.column_dimensions['D'].width = 20  # 出版社
    ws.column_dimensions['E'].width = 8   # 评分
    ws.column_dimensions['F'].width = 50  # 链接
    ws.column_dimensions['G'].width = 15  # 封面
    
    # 设置标题行高
    ws.row_dimensions[1].height = 20
    
    print("开始处理书籍数据并插入封面...")
    
    # 处理每本书
    for i, book in enumerate(books_data, 2):
        print(f"处理第 {i-1} 本书: {book['书名']}")
        
        # 填充基本信息
        ws.cell(row=i, column=1, value=i-1)  # 序号
        ws.cell(row=i, column=2, value=book['书名'])  # 书名
        ws.cell(row=i, column=3, value=book['作者'])  # 作者
        ws.cell(row=i, column=4, value=book['出版社'])  # 出版社
        ws.cell(row=i, column=5, value=book['评分'])  # 评分
        ws.cell(row=i, column=6, value=book['书籍链接'])  # 链接
        
        # 处理封面
        if book['封面链接']:
            try:
                # 查找对应的封面文件
                cover_files = [f for f in os.listdir('book_covers') if f.startswith(f"cover_{i-1}_")]
                
                if cover_files:
                    cover_path = f"book_covers/{cover_files[0]}"
                    
                    # 调整图片大小
                    img = PILImage.open(cover_path)
                    img.thumbnail((100, 140), PILImage.Resampling.LANCZOS)
                    
                    # 保存调整后的图片到临时文件
                    temp_path = f"temp_cover_{i-1}.jpg"
                    img.save(temp_path, 'JPEG', quality=85)
                    
                    # 插入到Excel
                    excel_img = Image(temp_path)
                    excel_img.width = 80
                    excel_img.height = 112
                    
                    # 设置图片位置
                    cell_ref = f'G{i}'
                    ws.add_image(excel_img, cell_ref)
                    
                    # 设置行高
                    ws.row_dimensions[i].height = 90
                    
                    # 删除临时文件
                    os.remove(temp_path)
                    print(f"  ✓ 封面已添加")
                else:
                    ws.cell(row=i, column=7, value="封面文件未找到")
                    print(f"  ✗ 封面文件未找到")
                    
            except Exception as e:
                ws.cell(row=i, column=7, value=f"封面处理失败: {str(e)}")
                print(f"  ✗ 封面处理失败: {e}")
        else:
            ws.cell(row=i, column=7, value="无封面")
            print(f"  - 无封面链接")
        
        # 设置文本对齐
        for col in range(1, 7):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='left', vertical='center')
    
    # 保存Excel文件
    excel_file = 'douban_books_sample_with_covers.xlsx'
    wb.save(excel_file)
    print(f"\nExcel文件已保存: {excel_file}")
    
    return excel_file

def main():
    """主函数"""
    try:
        print("=== 创建完整数据Excel文件 ===")
        excel_file1 = create_excel_from_data()
        
        print("\n=== 创建带封面示例Excel文件 ===")
        excel_file2 = create_excel_with_images()
        
        print(f"\n=== 完成 ===")
        print(f"✓ 完整数据Excel文件: {excel_file1}")
        print(f"✓ 带封面示例Excel文件: {excel_file2}")
        print(f"✓ 两个文件都可以用Excel打开查看")
        
    except Exception as e:
        print(f"生成Excel文件时出错: {e}")

if __name__ == "__main__":
    main()
